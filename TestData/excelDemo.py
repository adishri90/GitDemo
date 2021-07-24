import openpyxl

book = openpyxl.load_workbook("C:\\Users\\adashriv\\Downloads\\PythonSelFramework\\TestData\\PythonExcelDemo.xlsx")
sheet = book.active
# Read value from excel
cell = sheet.cell(row=1, column=2)
print(cell.value)
# Write value to excel
sheet.cell(row=2, column=2).value = "Adarsh"
print(sheet.cell(row=2, column=2).value)
# Get total rows and columns
print(sheet.max_row)
print(sheet.max_column)
# Another way to get cell value
print(sheet['A3'].value)
# use for loop to print values
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
